from openpyxl.styles import PatternFill


def get_no_record_fill():
    return PatternFill('solid', fgColor='C00000')


def get_late_record_fill():
    return PatternFill('solid', fgColor='FFA500')
