import calendar
import time

import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import FillTools
import Times


# 创建表头
def create_header(sheet, cols):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    sheet.cell(1, 1).value = '刷卡记录表'
    # 样式
    font = Font('宋体', size=20, bold=True, italic=False, strike=False, color='008000')
    sheet.cell(1, 1).font = font
    alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(1, 1).alignment = alignment

    border = Border(left=Side(border_style='thick', color='000000'),
                    right=Side(border_style='thick', color='000000'),
                    top=Side(border_style='thick', color='000000'),
                    bottom=Side(border_style='thick', color='000000'))
    for col in range(1, cols + 1):
        sheet.cell(1, col).border = border


# 创建日期栏
def create_time_cells(sheet, times, cols):
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    sheet.cell(2, 1).value = '考勤日期'

    sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=12)
    duration = f'{times.year}/{times.start_month}/{times.start_day} ~ {times.end_month}/{times.end_day}'
    sheet.cell(2, 3, duration)

    sheet.merge_cells(start_row=2, start_column=cols - 4, end_row=2, end_column=cols - 3)
    sheet.cell(2, cols - 4).value = '制表时间:'

    sheet.merge_cells(start_row=2, start_column=cols - 2, end_row=2, end_column=cols)
    cur_time = time.strftime('%Y/%m/%d', time.localtime(time.time()))
    sheet.cell(2, cols - 2).value = cur_time


def create_week_cells(sheet, month_range, cols):
    days = month_range[1]
    week_of_first_day = month_range[0]
    weekends = []
    font = Font('宋体', size=12, bold=False, italic=False, strike=False, color='000080')
    fill = PatternFill('solid', fgColor='32CD32')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for day in range(0, days):
        week = (day + week_of_first_day) % 7
        if week in [5, 6]:
            # 周末
            sheet.cell(3, day + 2).fill = fill
            weekends.append(day)
        else:
            # 工作日
            pass
        sheet.cell(3, day + 2, day + 1)
        sheet.cell(3, day + 2).font = font
        sheet.cell(3, day + 2).alignment = alignment

    sheet.cell(3, cols, '签名')
    sheet.cell(3, cols).font = font
    sheet.cell(3, cols).alignment = alignment
    small_font = Font('宋体', size=8, bold=False, italic=False, strike=False, color='000080')
    sheet.cell(3, cols - 1, '60分钟<迟到≤120分钟每次扣除100元')
    sheet.cell(3, cols - 1).font = small_font
    sheet.cell(3, cols - 1).alignment = alignment
    sheet.cell(3, cols - 2, '30分钟<迟到≤60分钟每次扣除50元')
    sheet.cell(3, cols - 2).font = small_font
    sheet.cell(3, cols - 2).alignment = alignment
    sheet.cell(3, cols - 3, '迟到≤30分钟每次扣除30元')
    sheet.cell(3, cols - 3).font = small_font
    sheet.cell(3, cols - 3).alignment = alignment
    sheet.cell(3, cols - 4, '迟到≤10分钟+缺卡')
    sheet.cell(3, cols - 4).font = small_font
    sheet.cell(3, cols - 4).alignment = alignment
    sheet.row_dimensions[3].height = 32

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    for day in range(1, cols + 1):
        if day <= days + 1:
            sheet.column_dimensions[Times.get_column_string(day)].width = 6
            print(Times.get_column_string(day))
        sheet.cell(3, day).border = border

    return weekends


# 填充表内容
def create_record(record, row, weekends, sheet, cols):
    sheet.cell(row, 2, '工号')
    sheet.cell(row, 3, record.number)
    sheet.cell(row, 6, '姓名')
    sheet.cell(row, 7, record.name)

    sheet.merge_cells(start_row=row + 1, start_column=cols, end_row=row + 2, end_column=cols)
    sheet.merge_cells(start_row=row + 1, start_column=cols - 1, end_row=row + 2, end_column=cols - 1)
    sheet.merge_cells(start_row=row + 1, start_column=cols - 2, end_row=row + 2, end_column=cols - 2)
    sheet.merge_cells(start_row=row + 1, start_column=cols - 3, end_row=row + 2, end_column=cols - 3)
    sheet.merge_cells(start_row=row + 1, start_column=cols - 4, end_row=row + 2, end_column=cols - 4)

    late_less_ten_minutes_or_no_record_times = 0
    late_less_half_hour_times = 0
    late_less_one_hour_times = 0
    late_less_two_hour_times = 0
    check_in_minutes = 8 * 60 + 30
    check_out_minutes = 17 * 60 + 30

    # 样式
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for index in range(1, cols + 1):
        sheet.cell(row + 1, index).alignment = alignment
        sheet.cell(row + 2, index).alignment = alignment
        sheet.cell(row + 1, index).border = border
        sheet.cell(row + 2, index).border = border

    for index in range(0, len(record.records)):
        sheet.cell(row + 1, 1, '签到')
        sheet.cell(row + 2, 1, '签退')

        check_in_time = record.records[index].check_in
        check_out_time = record.records[index].check_out
        if index not in weekends:
            check_in_cell = sheet.cell(row + 1, index + 2)
            check_out_cell = sheet.cell(row + 2, index + 2)
            if len(check_in_time) == 0:
                # sheet.cell(row + 1, index + 2, "缺卡")
                # sheet.cell(row + 1, index + 2).fill = FillTools.get_no_record_fill()
                check_in_cell.value = '缺卡'
                check_in_cell.fill = FillTools.get_no_record_fill()
                late_less_ten_minutes_or_no_record_times += 1
            else:
                # sheet.cell(row + 1, index + 2, check_in_time)
                check_in_cell.value = check_in_time
                minute_string = check_in_time.split(':')
                minutes = int(minute_string[0]) * 60 + int(minute_string[1])
                # print(f'real check_in={minutes}')
                if minutes - check_in_minutes <= 0:
                    # 正常签到
                    pass
                elif minutes - check_in_minutes <= 10:
                    # 迟到10分钟以内
                    late_less_ten_minutes_or_no_record_times += 1
                    check_in_cell.fill = FillTools.get_late_record_fill()
                elif minutes - check_in_minutes <= 30:
                    # 迟到30分钟以内
                    late_less_half_hour_times += 1
                    check_in_cell.fill = FillTools.get_late_record_fill()
                elif minutes - check_in_minutes <= 60:
                    # 迟到1小时以内
                    late_less_one_hour_times += 1
                    check_in_cell.fill = FillTools.get_late_record_fill()
                else:
                    # 迟到大于1小时
                    late_less_two_hour_times += 1
                    check_in_cell.fill = FillTools.get_late_record_fill()

            if len(check_out_time) == 0:
                # sheet.cell(row + 2, index + 2, "缺卡")
                # sheet.cell(row + 2, index + 2).fill = FillTools.get_no_record_fill()
                check_out_cell.value = '缺卡'
                check_out_cell.fill = FillTools.get_no_record_fill()
                late_less_ten_minutes_or_no_record_times += 1
            else:
                # sheet.cell(row + 2, index + 2, check_out_time)
                check_out_cell.value = check_out_time
                minute_string = check_out_time.split(':')
                minutes = int(minute_string[0]) * 60 + int(minute_string[1])
                # print(f'real check_out={minutes}')
                if check_out_minutes - minutes <= 0:
                    # 正常签退
                    pass
                elif check_out_minutes - minutes <= 10:
                    # 早退10分钟以内
                    late_less_ten_minutes_or_no_record_times += 1
                    check_out_cell.fill = FillTools.get_late_record_fill()
                elif check_out_minutes - minutes <= 30:
                    # 早退30分钟以内
                    late_less_half_hour_times += 1
                    check_out_cell.fill = FillTools.get_late_record_fill()
                elif check_out_minutes - minutes <= 60:
                    # 早退1小时以内
                    late_less_one_hour_times += 1
                    check_out_cell.fill = FillTools.get_late_record_fill()
                else:
                    # 早退1小时以上
                    late_less_two_hour_times += 1
                    check_out_cell.fill = FillTools.get_late_record_fill()

        sheet.cell(row + 1, cols - 4, late_less_ten_minutes_or_no_record_times)
        sheet.cell(row + 1, cols - 3, late_less_half_hour_times)
        sheet.cell(row + 1, cols - 2, late_less_one_hour_times)
        sheet.cell(row + 1, cols - 1, late_less_two_hour_times)


def save_records_to_file(records):
    month_range = calendar.monthrange(records.time.year, records.time.start_month)
    total_cols = month_range[1] + 6
    print(month_range[0], month_range[1])
    wb = xl.Workbook()
    ws = wb.active
    create_header(ws, total_cols)
    create_time_cells(ws, records.time, total_cols)
    weekends = create_week_cells(ws, month_range, total_cols)
    print(weekends)
    for index in range(0, len(records.persons)):
        create_record(records.persons[index], index * 3 + 4, weekends, ws, total_cols)

    times = str(
        f'{records.time.year}/{records.time.start_month}/{records.time.start_day} - {records.time.end_month}/{records.time.end_day}')
    wb.save('generate.xlsx')
