import openpyxl as xl
from openpyxl.chart import BarChart, Reference

from Calculate import calculate_records
from CreateFile import save_records_to_file
#
# wb = xl.load_workbook('test.xlsx')
# sheet = wb['Sheet1']
# cell = sheet.cell(1, 1)
# for row in range(2, sheet.max_row + 1):
#     value = sheet.cell(row, 2)
#     discount = sheet.cell(row, 3)
#     price = sheet.cell(row, 4)
#     price.value = value.value * discount.value
#
# values = Reference(sheet,
#                    min_row=2,
#                    max_row=sheet.max_row + 1,
#                    min_col=3,
#                    max_col=3)
#
# barChart = BarChart()
# barChart.add_data(values)
# sheet.add_chart(barChart, "e2")
#
# wb.save("new_test.xlsx")

# Numpy Pandas MatPlotLib Scikit-Learn

records = calculate_records('12月汇总表.xls')
save_records_to_file(records)

